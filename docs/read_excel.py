"""Script đọc file CTDT_K50.xlsx và xuất ra JSON để phân tích"""
import openpyxl
import json

wb = openpyxl.load_workbook(r'd:\PROJECT\CT553_LVTN\docs\CTDT_K50.xlsx', data_only=True)

result = {"sheets": wb.sheetnames, "data": {}}

for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append([str(c) if c is not None else None for c in row])
    result["data"][name] = {
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "rows": rows
    }

with open(r'd:\PROJECT\CT553_LVTN\docs\excel_parsed.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print("Done! Saved to excel_parsed.json")
print(f"Sheets: {wb.sheetnames}")
for name in wb.sheetnames:
    print(f"  {name}: {len(result['data'][name]['rows'])} rows, {result['data'][name]['max_col']} cols")
