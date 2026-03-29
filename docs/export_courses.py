import openpyxl
import json

wb = openpyxl.load_workbook('docs/CTDT_K50.xlsx')
ws1 = wb[wb.sheetnames[0]]

courses = []
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    ma = row[1].value
    if ma:
        course = {
            'code': str(ma).strip(),
            'name': str(row[2].value).strip() if row[2].value else '',
            'credits': int(row[3].value) if row[3].value and str(row[3].value).isdigit() else str(row[3].value).strip() if row[3].value else 0,
            'required': str(row[4].value).strip() if row[4].value else '',
            'elective': str(row[5].value).strip() if row[5].value else '',
            'prerequisite': str(row[6].value).strip() if row[6].value else '',
            'corequisite': str(row[7].value).strip() if row[7].value else '',
            'description': str(row[8].value).strip() if row[8].value else '',
            'theory': str(row[9].value).strip() if row[9].value else '',
            'practice': str(row[10].value).strip() if row[10].value else '',
        }
        courses.append(course)

with open('docs/all_courses_full.json', 'w', encoding='utf-8') as f:
    json.dump(courses, f, ensure_ascii=False, indent=2)

print("Exported {} courses".format(len(courses)))

# Also export curriculum plan (sheet 2)
ws2 = wb['KyThuatPhanMem']
semesters = []
current_semester = None

for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row):
    c0 = str(row[0].value).strip() if row[0].value is not None else ''
    c1 = str(row[1].value).strip() if row[1].value is not None else ''
    
    if c0.startswith('Hoc ky') or c0.startswith('Học kỳ'):
        current_semester = {'name': c0, 'courses': []}
        semesters.append(current_semester)
    elif c1 and current_semester is not None:
        current_semester['courses'].append(c1)

with open('docs/curriculum_plan.json', 'w', encoding='utf-8') as f:
    json.dump(semesters, f, ensure_ascii=False, indent=2)

print("Exported {} semesters".format(len(semesters)))
for s in semesters:
    print("  {} -> {} courses: {}".format(s['name'], len(s['courses']), ', '.join(s['courses'])))
